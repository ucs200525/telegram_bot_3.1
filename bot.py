import platform
import pandas as pd
import requests
from datetime import datetime, timedelta
from telegram import Update
from telegram.ext import Application, CommandHandler, CallbackContext, ConversationHandler, MessageHandler, filters
import pytz
from openpyxl import load_workbook
from opencage.geocoder import OpenCageGeocode
import os
import logging
import subprocess
from flask import Flask

# Configure logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Define conversation states
LOCATION = 0
DATE_LOCATION = 1
CONFIRMATION = 2

# Function to get sunrise and sunset times for a specific date
def get_sun_times_for_date(lat, lng, local_tz, date):
    next_day = (date + timedelta(days=1)).strftime('%Y-%m-%d')
    date_str = date.strftime('%Y-%m-%d')

    url = f'https://api.sunrise-sunset.org/json?lat={lat}&lng={lng}&formatted=0&date='

    date_response = requests.get(url + date_str).json()
    next_day_response = requests.get(url + next_day).json()

    sunrise_date_utc = date_response['results']['sunrise']
    sunset_date_utc = date_response['results']['sunset']
    sunrise_next_day_utc = next_day_response['results']['sunrise']

    # Convert to local time
    ist = pytz.timezone(local_tz)
    sunrise_date = pd.to_datetime(sunrise_date_utc).tz_convert(ist)
    sunset_date = pd.to_datetime(sunset_date_utc).tz_convert(ist)
    sunrise_next_day = pd.to_datetime(sunrise_next_day_utc).tz_convert(ist)

    return sunrise_date, sunset_date, sunrise_next_day

# Function to update Excel file with specific date's sunrise and sunset times
def update_excel_with_date(file_path, sunrise_date, sunset_date, sunrise_next_day, day_of_week):
    try:
        wb = load_workbook(file_path)
        sheet_names = wb.sheetnames
        today_sheet_name = day_of_week.upper()

        if today_sheet_name in sheet_names:
            ws = wb[today_sheet_name]
            ws['O4'] = sunrise_date.strftime('%H:%M:%S')
            ws['O5'] = sunset_date.strftime('%H:%M:%S')
            ws['O6'] = sunrise_next_day.strftime('%H:%M:%S')

            wb.save(file_path)
            wb.close()
            logger.info(f'Excel file {file_path} updated successfully with sheet {today_sheet_name}.')
        else:
            logger.error(f'Sheet {today_sheet_name} not found in Excel file.')
    except Exception as e:
        logger.error(f'Error updating Excel file: {e}')

# Function to convert Excel to image using Node.js script with specific date
def save_excel_as_image_with_nodejs_date(excel_file_path, output_image_path, sheet_name, date_str):
    try:
        node_script = r's/image2.js'  # Update this path to your Node.js script
        subprocess.run(['node', node_script, excel_file_path, sheet_name, output_image_path, date_str], check=True)
        logger.info(f"Image successfully saved to {output_image_path}")
    except Exception as e:
        logger.error(f"Error converting Excel to image: {e}")

# Function to get DrikPanchang screenshot using Node.js script with specific date
def get_drikpanchang_screenshot_date(city, date, output_image_path):
    try:
        node_script = r'/newProj.js'  # Update this path to your Node.js script
        subprocess.run(['node', node_script, city, date, output_image_path], check=True)
        logger.info(f"DrikPanchang screenshot successfully saved to {output_image_path}")
    except Exception as e:
        logger.error(f"Error capturing DrikPanchang screenshot: {e}")

# Flask application
app = Flask(__name__)

@app.route('/start-bot', methods=['GET'])
def start_bot():
    main()
    return "Bot started", 200

# Command handler to start the conversation for /gt
async def send_table_start(update: Update, context: CallbackContext):
    await update.message.reply_text("Please enter your location (e.g., Vijayawada) or type /cancel to abort.")
    return LOCATION

# Command handler to start the conversation for /dgt
async def send_date_location_start(update: Update, context: CallbackContext):
    await update.message.reply_text("Please enter the date (dd/mm/yyyy) and location separated by a comma (e.g., 26/06/2024, Vijayawada) or type /cancel to abort.")
    return DATE_LOCATION

# Function to handle location input and send the table with images
async def receive_location(update: Update, context: CallbackContext):
    location = update.message.text
    user_id = update.message.from_user.id
    username = update.effective_user.username
    logger.info(f'User {user_id} , {username} sent location: {location}')
    await update.message.reply_text("Your task is in progress...")

    # Use OpenCage Geocoder to get coordinates
    geocoder = OpenCageGeocode(context.bot_data['opencage_api_key'])
    result = geocoder.geocode(location)

    if result and len(result):
        latitude = result[0]['geometry']['lat']
        longitude = result[0]['geometry']['lng']
        logger.info(f'Coordinates for {location}: {latitude}, {longitude}')
        
        local_tz = 'Asia/Kolkata'  # Assuming Indian Standard Time (IST)
        
        # Get sun times for today
        today = datetime.now()
        day_of_week = today.strftime('%A').upper()  # Get day of the week
        
        sunrise_today, sunset_today, sunrise_tomorrow = get_sun_times_for_date(latitude, longitude, local_tz, today)

        # Update the Excel file with today's sheet
        file_path = context.bot_data['excel_file_path']
        update_excel_with_date(file_path, sunrise_today, sunset_today, sunrise_tomorrow, day_of_week)

        # Save Excel range as image
        save_image_path = context.bot_data['image_save_path']
        save_excel_as_image_with_nodejs_date(file_path, save_image_path, day_of_week, today.strftime('%d/%m/%Y'))

        # Get DrikPanchang screenshot
        drikpanchang_image_path = context.bot_data['drikpanchang_image_path']
        city = location  # Use the provided location
        date = today.strftime('%d/%m/%Y')  # Current date in dd/mm/yyyy format
        get_drikpanchang_screenshot_date(city, date, drikpanchang_image_path)

        # Notify user of the day of the week
        await update.message.reply_text(f"Day of the week for today ({today.strftime('%d/%m/%Y')}) is {day_of_week}")

        # Ask user if they want the Excel sheet or an image of the Excel sheet
        await update.message.reply_text("Do you want the Excel sheet or an image of the Excel sheet? Reply with 'Excel' or 'Image'.")
        context.user_data['file_path'] = file_path
        context.user_data['save_image_path'] = save_image_path
        context.user_data['drikpanchang_image_path'] = drikpanchang_image_path
        return CONFIRMATION
    else:
        logger.warning(f'Could not find coordinates for location: {location}')
        await update.message.reply_text("Sorry, I couldn't find coordinates for that location. Please try again.")
        return LOCATION

# Function to handle date and location input
async def receive_date_location(update: Update, context: CallbackContext):
    date_location = update.message.text.split(',')
    if len(date_location) != 2:
        await update.message.reply_text("Invalid input. Please enter the date and location separated by a comma (e.g., 26/06/2024, Vijayawada) or type /cancel to abort.")
        return DATE_LOCATION

    date_str, location = date_location[0].strip(), date_location[1].strip()
    user_id = update.message.from_user.id
    username = update.effective_user.username
    logger.info(f'User {user_id} , {username} sent date: {date_str} and location: {location}')
    await update.message.reply_text("Your task is in progress...")

    try:
        date = datetime.strptime(date_str, '%d/%m/%Y')
    except ValueError:
        await update.message.reply_text("Invalid date format. Please enter the date in dd/mm/yyyy format.")
        return DATE_LOCATION

    # Use OpenCage Geocoder to get coordinates
    geocoder = OpenCageGeocode(context.bot_data['opencage_api_key'])
    result = geocoder.geocode(location)

    if result and len(result):
        latitude = result[0]['geometry']['lat']
        longitude = result[0]['geometry']['lng']
        logger.info(f'Coordinates for {location}: {latitude}, {longitude}')
        
        local_tz = 'Asia/Kolkata'  # Assuming Indian Standard Time (IST)
        
        # Get day of the week for the given date
        day_of_week = date.strftime('%A').upper()

        # Get sun times for the given date
        sunrise_date, sunset_date, sunrise_next_day = get_sun_times_for_date(latitude, longitude, local_tz, date)

        # Update the Excel file with the specific date's sheet
        file_path = context.bot_data['excel_file_path']
        update_excel_with_date(file_path, sunrise_date, sunset_date, sunrise_next_day, day_of_week)

        # Save Excel range as image
        save_image_path = context.bot_data['image_save_path']
        save_excel_as_image_with_nodejs_date(file_path, save_image_path, day_of_week, date.strftime('%d/%m/%Y'))

        # Get DrikPanchang screenshot
        drikpanchang_image_path = context.bot_data['drikpanchang_image_path']
        city = location  # Use the provided location
        get_drikpanchang_screenshot_date(city, date.strftime('%d/%m/%Y'), drikpanchang_image_path)

        # Notify user of the day of the week
        await update.message.reply_text(f"Day of the week for {date.strftime('%d/%m/%Y')} is {day_of_week}")

        # Ask user if they want the Excel sheet or an image of the Excel sheet
        await update.message.reply_text("Do you want the Excel sheet or an image of the Excel sheet? Reply with 'Excel' or 'Image'.")
        context.user_data['file_path'] = file_path
        context.user_data['save_image_path'] = save_image_path
        context.user_data['drikpanchang_image_path'] = drikpanchang_image_path
        return CONFIRMATION
    else:
        logger.warning(f'Could not find coordinates for location: {location}')
        await update.message.reply_text("Sorry, I couldn't find coordinates for that location. Please try again.")
        return DATE_LOCATION

# Function to handle confirmation from the user
async def handle_confirmation(update: Update, context: CallbackContext):
    choice = update.message.text.lower()
    file_path = context.user_data.get('file_path')
    save_image_path = context.user_data.get('save_image_path')
    drikpanchang_image_path = context.user_data.get('drikpanchang_image_path')

    if choice == 'excel':
        await update.message.reply_document(document=open(file_path, 'rb'))
    elif choice == 'image':
        await update.message.reply_photo(photo=open(save_image_path, 'rb'))
    else:
        await update.message.reply_text("Invalid choice. Please reply with 'Excel' or 'Image'.")
        return CONFIRMATION

    # Send DrikPanchang screenshot
    await update.message.reply_photo(photo=open(drikpanchang_image_path, 'rb'))

    # End the conversation
    return ConversationHandler.END

# Command handler for /start
async def start_command_handler(update: Update, context: CallbackContext):
    start_message = (
        "Welcome to the Panchangam Bot!\n"
        "Here are the available commands:\n\n"
        "/gt - Get a screenshot of Bhargava Panchangam and Drik Panchangam for your location. Just send your location (e.g., Vijayawada).\n"
        "/dgt - Get a screenshot of Bhargava Panchangam and Drik Panchangam for a specific date and location. Send the date and location separated by a comma (e.g., 26/06/2024, Vijayawada).\n"
        "/cancel - Cancel the current operation."
    )
    await update.message.reply_text(start_message)

async def help_command_handler(update: Update, context: CallbackContext):
    await update.message.reply_text("This is the help message.")

async def main_handler(update: Update, context: CallbackContext):
    await update.message.reply_text("You sent a text message.")

# Command handler for handling unknown commands
async def unknown_command_handler(update: Update, context: CallbackContext):
    await update.message.reply_text("Sorry, I didn't understand that command. Please use one of the available commands:\n"
                                      "/gt - Get a screenshot of Bhargava Panchangam and Drik Panchangam for your location. Just send your location (e.g., Vijayawada).\n"
                                      "/dgt - Get a screenshot of Bhargava Panchangam and Drik Panchangam for a specific date and location. Send the date and location separated by a comma (e.g., 26/06/2024, Vijayawada).\n"
                                      "/cancel - Cancel the current operation.")

# Command handler to cancel ongoing operations
async def cancel_command_handler(update: Update, context: CallbackContext):
    await update.message.reply_text("Operation cancelled.")
    return ConversationHandler.END  # End any active conversation

def main():
    # Prompt user to enter tokens and paths
    opencage_api_key = '699522e909454a09b82d1c728fc79925'
    excel_file_path = r'/Bharghava_Siddhanta_Panchangam.xlsx'
    image_save_path = r'/ExcelToImage.png'
    drikpanchang_image_path = r'/DrikPanchangImage.png'
    bot_token = '7274941037:AAHIWiU5yvfIzo7eJWPu9S5CeJIid6ATEyM'

    # Create the Application instance
    application = Application.builder().token(bot_token).build()

    # Save tokens and paths in bot_data
    application.bot_data['opencage_api_key'] = opencage_api_key
    application.bot_data['excel_file_path'] = excel_file_path
    application.bot_data['image_save_path'] = image_save_path
    application.bot_data['drikpanchang_image_path'] = drikpanchang_image_path

    # Create the conversation handler for /gt
    gt_conversation_handler = ConversationHandler(
        entry_points=[CommandHandler('gt', send_table_start)],
        states={
            LOCATION: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_location)],
            CONFIRMATION: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_confirmation)],
        },
        fallbacks=[CommandHandler('cancel', cancel_command_handler)]
    )

    # Create the conversation handler for /dgt
    dgt_conversation_handler = ConversationHandler(
        entry_points=[CommandHandler('dgt', send_date_location_start)],
        states={
            DATE_LOCATION: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_date_location)],
            CONFIRMATION: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_confirmation)],
        },
        fallbacks=[CommandHandler('cancel', cancel_command_handler)]
    )

    # Add conversation handlers
    application.add_handler(gt_conversation_handler)
    application.add_handler(dgt_conversation_handler)

    # Add other handlers
    application.add_handler(CommandHandler("start", start_command_handler))
    application.add_handler(CommandHandler("help", help_command_handler))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, unknown_command_handler))
    application.add_handler(MessageHandler(filters.TEXT, main_handler))

    # Run the bot
    application.run_polling()
    logger.info('Bot started successfully.')

if __name__ == '__main__':
    main()
